from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
from werkzeug.utils import secure_filename
import pandas as pd
import os
from datetime import datetime
from src.document_parser import DocumentParser
from src.recommendation_pipeline import CandidateRecommendationPipeline
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max for multiple files
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DATABASE_FOLDER'] = 'database'
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'txt'}

# Ensure folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['DATABASE_FOLDER'], exist_ok=True)

# Initialize parsers
document_parser = DocumentParser()

# Initialize recommendation pipeline
try:
    pipeline = CandidateRecommendationPipeline(vectorizer_path='models/vectorizer.pkl')
    print("✓ Recommendation pipeline initialized")
except Exception as e:
    print(f"✗ Pipeline initialization error: {str(e)}")
    pipeline = None

# Global storage
cvs_data = []
jobs_data = []


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_cvs_to_excel():
    """Save all CVs data to Excel database"""
    try:
        if not cvs_data:
            return
        
        # Prepare data for Excel
        excel_data = []
        for cv in cvs_data:
            excel_data.append({
                'Candidate ID': cv['candidate_id'],
                'Name': cv['name'],
                'Skills': cv['skills'][:500] if cv['skills'] else '',
                'Experience': cv['experience'][:500] if cv['experience'] else '',
                'Education': cv['education'][:300] if cv['education'] else '',
                'Filename': cv['filename'],
                'Upload Date': cv['timestamp'],
                'Text Length': len(cv.get('cv_text', ''))
            })
        
        df = pd.DataFrame(excel_data)
        excel_path = os.path.join(app.config['DATABASE_FOLDER'], 'cvs_database.xlsx')
        
        # Save to Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='CVs', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['CVs']
            
            # Header formatting
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 50
            worksheet.column_dimensions['D'].width = 50
            worksheet.column_dimensions['E'].width = 35
            worksheet.column_dimensions['F'].width = 30
            worksheet.column_dimensions['G'].width = 20
            worksheet.column_dimensions['H'].width = 12
        
        print(f"✓ CVs database saved: {excel_path}")
    except Exception as e:
        print(f"✗ Error saving CVs to Excel: {str(e)}")


def save_jobs_to_excel():
    """Save all Jobs data to Excel database"""
    try:
        if not jobs_data:
            return
        
        # Prepare data for Excel
        excel_data = []
        for job in jobs_data:
            excel_data.append({
                'Job ID': job['job_id'],
                'Title': job['title'],
                'Required Skills': job['required_skills'][:500] if job['required_skills'] else '',
                'Experience Required': job['experience_required'][:500] if job['experience_required'] else '',
                'Education Required': job['education_required'][:300] if job['education_required'] else '',
                'Source': job.get('source', 'file'),
                'Filename': job.get('filename', 'N/A'),
                'Upload Date': job['timestamp'],
                'Description Length': len(job.get('job_description', ''))
            })
        
        df = pd.DataFrame(excel_data)
        excel_path = os.path.join(app.config['DATABASE_FOLDER'], 'jobs_database.xlsx')
        
        # Save to Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Jobs', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Jobs']
            
            # Header formatting
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 50
            worksheet.column_dimensions['D'].width = 50
            worksheet.column_dimensions['E'].width = 35
            worksheet.column_dimensions['F'].width = 12
            worksheet.column_dimensions['G'].width = 30
            worksheet.column_dimensions['H'].width = 20
            worksheet.column_dimensions['I'].width = 15
        
        print(f"✓ Jobs database saved: {excel_path}")
    except Exception as e:
        print(f"✗ Error saving Jobs to Excel: {str(e)}")


def save_recommendations_to_excel(recommendations_data):
    """Save recommendations/matches data to Excel database"""
    try:
        if not recommendations_data:
            return
        
        # Prepare data for Excel
        excel_data = []
        for job_rec in recommendations_data:
            job_id = job_rec['job_id']
            job_title = job_rec['job_title']
            
            for candidate in job_rec['candidates']:
                excel_data.append({
                    'Job ID': job_id,
                    'Job Title': job_title,
                    'Rank': candidate['rank'],
                    'Candidate ID': candidate['candidate_id'],
                    'Candidate Name': candidate['name'],
                    'Match %': candidate['match_percentage'],
                    'Similarity Score': candidate['similarity_score'],
                    'Summary': candidate['summary'][:300] if candidate['summary'] else '',
                    'Skills Match': candidate['skills'][:200] if candidate['skills'] else '',
                    'Generated Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
        
        df = pd.DataFrame(excel_data)
        excel_path = os.path.join(app.config['DATABASE_FOLDER'], 'recommendations_database.xlsx')
        
        # Check if file exists to append or create new
        if os.path.exists(excel_path):
            # Append to existing file
            existing_df = pd.read_excel(excel_path, sheet_name='Recommendations')
            df = pd.concat([existing_df, df], ignore_index=True)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Recommendations', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Recommendations']
            
            # Header formatting
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 8
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 25
            worksheet.column_dimensions['F'].width = 10
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 40
            worksheet.column_dimensions['I'].width = 30
            worksheet.column_dimensions['J'].width = 20
        
        print(f"✓ Recommendations database saved: {excel_path}")
    except Exception as e:
        print(f"✗ Error saving Recommendations to Excel: {str(e)}")


@app.route('/')
def index():
    """Serve the main page"""
    return render_template('index.html')


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'pipeline_loaded': pipeline is not None,
        'supported_formats': ['pdf', 'docx', 'txt']
    })


@app.route('/api/upload-cvs-bulk', methods=['POST'])
def upload_cvs_bulk():
    """
    Upload multiple CV files at once
    Automatically generates IDs and extracts names from filenames
    """
    try:
        if 'files[]' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files[]')
        
        if not files or len(files) == 0:
            return jsonify({'error': 'No files selected'}), 400
        
        successful_uploads = []
        failed_uploads = []
        
        for file in files:
            if file.filename == '':
                continue
                
            if not allowed_file(file.filename):
                failed_uploads.append({
                    'filename': file.filename,
                    'error': 'Invalid file format'
                })
                continue
            
            try:
                # Generate unique ID
                candidate_id = str(uuid.uuid4())[:8]
                
                # Extract name from filename (remove extension)
                name = os.path.splitext(file.filename)[0]
                
                # Save file
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_filename = f"cv_{candidate_id}_{timestamp}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(file_path)
                
                # Parse document
                extracted_text = document_parser.parse_file(file_path)
                sections = document_parser.extract_sections(extracted_text)
                
                # Create CV entry
                cv_entry = {
                    'candidate_id': candidate_id,
                    'name': name,
                    'skills': sections.get('skills', ''),
                    'experience': sections.get('experience', ''),
                    'education': sections.get('education', ''),
                    'cv_text': sections.get('full_text', ''),
                    'filename': filename,
                    'upload_path': file_path,
                    'timestamp': datetime.now().isoformat()
                }
                
                # Process with pipeline
                if pipeline:
                    combined_text = f"{cv_entry['skills']} {cv_entry['experience']} {cv_entry['education']} {cv_entry['cv_text']}"
                    cv_entry['cleaned_text'] = pipeline.clean_text(combined_text)
                
                # Store
                cvs_data.append(cv_entry)
                
                successful_uploads.append({
                    'filename': filename,
                    'candidate_id': candidate_id,
                    'name': name,
                    'text_length': len(extracted_text)
                })
                
            except Exception as e:
                failed_uploads.append({
                    'filename': file.filename,
                    'error': str(e)
                })
                if os.path.exists(file_path):
                    os.remove(file_path)
        
        # Save CVs to Excel database
        save_cvs_to_excel()
        
        return jsonify({
            'success': True,
            'message': f'Processed {len(successful_uploads)} CVs successfully',
            'successful': successful_uploads,
            'failed': failed_uploads,
            'total_cvs': len(cvs_data)
        }), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload-job-file', methods=['POST'])
def upload_job_file():
    """
    Upload Job Description as document file (PDF, DOCX, TXT)
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file format. Use PDF, DOCX, or TXT'}), 400
        
        # Generate unique ID
        job_id = str(uuid.uuid4())[:8]
        
        # Extract title from filename
        title = os.path.splitext(file.filename)[0]
        
        # Save file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"job_{job_id}_{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)
        
        # Parse document
        try:
            extracted_text = document_parser.parse_file(file_path)
            sections = document_parser.extract_sections(extracted_text)
        except Exception as e:
            os.remove(file_path)
            return jsonify({'error': f'Failed to parse document: {str(e)}'}), 500
        
        # Create job entry
        job_entry = {
            'job_id': job_id,
            'title': title,
            'required_skills': sections.get('skills', ''),
            'experience_required': sections.get('experience', ''),
            'education_required': sections.get('education', ''),
            'job_description': sections.get('full_text', ''),
            'filename': filename,
            'upload_path': file_path,
            'timestamp': datetime.now().isoformat(),
            'source': 'file'
        }
        
        # Process with pipeline
        if pipeline:
            combined_text = f"{job_entry['required_skills']} {job_entry['experience_required']} {job_entry['education_required']} {job_entry['job_description']}"
            job_entry['cleaned_text'] = pipeline.clean_text(combined_text)
        
        # Store
        jobs_data.append(job_entry)
        
        # Save Jobs to Excel database
        save_jobs_to_excel()
        
        return jsonify({
            'success': True,
            'message': 'Job description uploaded successfully',
            'job_id': job_id,
            'title': title,
            'text_length': len(extracted_text),
            'total_jobs': len(jobs_data)
        }), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/add-job-text', methods=['POST'])
def add_job_text():
    """
    Add Job Description as plain text
    """
    try:
        data = request.json
        job_text = data.get('job_description', '').strip()
        title = data.get('title', 'Untitled Position').strip()
        
        if not job_text:
            return jsonify({'error': 'Job description text is required'}), 400
        
        # Generate unique ID
        job_id = str(uuid.uuid4())[:8]
        
        # Extract sections from text
        sections = document_parser.extract_sections(job_text)
        
        # Create job entry
        job_entry = {
            'job_id': job_id,
            'title': title,
            'required_skills': sections.get('skills', ''),
            'experience_required': sections.get('experience', ''),
            'education_required': sections.get('education', ''),
            'job_description': job_text,
            'filename': None,
            'upload_path': None,
            'timestamp': datetime.now().isoformat(),
            'source': 'text'
        }
        
        # Process with pipeline
        if pipeline:
            combined_text = f"{job_entry['required_skills']} {job_entry['experience_required']} {job_entry['education_required']} {job_entry['job_description']}"
            job_entry['cleaned_text'] = pipeline.clean_text(combined_text)
        
        # Store
        jobs_data.append(job_entry)
        
        # Save Jobs to Excel database
        save_jobs_to_excel()
        
        return jsonify({
            'success': True,
            'message': 'Job description added successfully',
            'job_id': job_id,
            'title': title,
            'text_length': len(job_text),
            'total_jobs': len(jobs_data)
        }), 201
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/recommend', methods=['POST'])
def recommend():
    """Generate ranked recommendations with summarization"""
    try:
        data = request.json
        job_id = data.get('job_id', None)
        
        if len(cvs_data) == 0:
            return jsonify({'error': 'No CVs uploaded yet'}), 400
        
        if len(jobs_data) == 0:
            return jsonify({'error': 'No jobs uploaded yet'}), 400
        
        if not pipeline:
            return jsonify({'error': 'Pipeline not initialized'}), 500
        
        # Convert to DataFrames
        cvs_df = pd.DataFrame(cvs_data)
        jobs_df = pd.DataFrame(jobs_data)
        
        # Filter for specific job if requested
        if job_id:
            jobs_df = jobs_df[jobs_df['job_id'] == job_id]
            if len(jobs_df) == 0:
                return jsonify({'error': f'Job ID {job_id} not found'}), 404
        
        # Get all recommendations (rank all CVs for each job)
        recommendations_df = pipeline.batch_recommend(jobs_df, cvs_df, top_n=len(cvs_df))
        
        # Group recommendations by job and format response
        job_recommendations = []
        
        for _, job in jobs_df.iterrows():
            # Get recommendations for this job
            job_recs = recommendations_df[recommendations_df['job_id'] == job['job_id']]
            
            # Sort by similarity score (already ranked by batch_recommend)
            job_recs = job_recs.sort_values('similarity_score', ascending=False).reset_index(drop=True)
            
            # Generate summary for each candidate
            candidates_list = []
            for idx, row in job_recs.iterrows():
                # Find the CV data
                cv = next((c for c in cvs_data if c['candidate_id'] == row['candidate_id']), None)
                
                # Get candidate name from CV data
                candidate_name = cv['name'] if cv else 'Unknown'
                
                summary = generate_candidate_summary(cv, row['similarity_score'])
                
                candidates_list.append({
                    'rank': idx + 1,
                    'candidate_id': row['candidate_id'],
                    'name': candidate_name,
                    'similarity_score': round(row['similarity_score'], 4),
                    'match_percentage': round(row['similarity_score'] * 100, 2),
                    'summary': summary,
                    'skills': cv['skills'][:300] if cv else '',
                    'experience': cv['experience'][:300] if cv else '',
                    'education': cv['education'][:200] if cv else ''
                })
            
            job_recommendations.append({
                'job_id': job['job_id'],
                'job_title': job['title'],
                'candidates': candidates_list,
                'total_matches': len(candidates_list)
            })
        
        # Save recommendations to Excel database
        save_recommendations_to_excel(job_recommendations)
        
        return jsonify({
            'success': True,
            'jobs': job_recommendations,
            'total_jobs': len(jobs_df),
            'total_candidates': len(cvs_data),
            'timestamp': datetime.now().isoformat()
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def generate_candidate_summary(cv, score):
    """Generate a brief summary of candidate suitability"""
    if not cv:
        return "No candidate data available"
    
    match_level = "Excellent" if score > 0.7 else "Good" if score > 0.5 else "Moderate" if score > 0.3 else "Low"
    
    skills_preview = cv['skills'][:100].strip() if cv['skills'] else "No skills data"
    exp_preview = cv['experience'][:100].strip() if cv['experience'] else "No experience data"
    
    summary = f"{match_level} match. "
    
    if skills_preview != "No skills data":
        summary += f"Key skills: {skills_preview}... "
    
    if exp_preview != "No experience data":
        summary += f"Experience: {exp_preview}..."
    
    return summary


@app.route('/api/cvs', methods=['GET'])
def get_cvs():
    """Get all uploaded CVs"""
    cv_list = [{
        'candidate_id': cv['candidate_id'],
        'name': cv['name'],
        'filename': cv['filename'],
        'timestamp': cv['timestamp']
    } for cv in cvs_data]
    
    return jsonify({
        'cvs': cv_list,
        'total': len(cvs_data)
    }), 200


@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    """Get all uploaded jobs"""
    job_list = [{
        'job_id': job['job_id'],
        'title': job['title'],
        'source': job.get('source', 'file'),
        'timestamp': job['timestamp']
    } for job in jobs_data]
    
    return jsonify({
        'jobs': job_list,
        'total': len(jobs_data)
    }), 200


@app.route('/api/clear', methods=['POST'])
def clear_data():
    """Clear all data and uploaded files"""
    global cvs_data, jobs_data
    
    # Delete uploaded files
    for cv in cvs_data:
        if 'upload_path' in cv and cv['upload_path'] and os.path.exists(cv['upload_path']):
            try:
                os.remove(cv['upload_path'])
            except:
                pass
    
    for job in jobs_data:
        if 'upload_path' in job and job['upload_path'] and os.path.exists(job['upload_path']):
            try:
                os.remove(job['upload_path'])
            except:
                pass
    
    # Delete Excel database files
    database_files = ['cvs_database.xlsx', 'jobs_database.xlsx', 'recommendations_database.xlsx']
    for db_file in database_files:
        db_path = os.path.join(app.config['DATABASE_FOLDER'], db_file)
        if os.path.exists(db_path):
            try:
                os.remove(db_path)
                print(f"✓ Deleted {db_file}")
            except:
                pass
    
    cvs_data = []
    jobs_data = []
    
    return jsonify({'message': 'All data cleared successfully'}), 200


@app.route('/api/database-status', methods=['GET'])
def get_database_status():
    """Get Excel database files status"""
    database_info = []
    database_files = {
        'cvs_database.xlsx': 'CVs Database',
        'jobs_database.xlsx': 'Jobs Database',
        'recommendations_database.xlsx': 'Recommendations Database'
    }
    
    for filename, description in database_files.items():
        file_path = os.path.join(app.config['DATABASE_FOLDER'], filename)
        if os.path.exists(file_path):
            file_stats = os.stat(file_path)
            database_info.append({
                'name': description,
                'filename': filename,
                'size': f"{file_stats.st_size / 1024:.2f} KB",
                'last_modified': datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'exists': True
            })
        else:
            database_info.append({
                'name': description,
                'filename': filename,
                'exists': False
            })
    
    return jsonify({
        'databases': database_info,
        'database_folder': app.config['DATABASE_FOLDER']
    }), 200


if __name__ == '__main__':
    print("\n" + "="*70)
    print("   🚀 Enhanced Candidate Recommendation System")
    print("="*70)
    print("\n📄 Supported Formats: PDF, DOCX, TXT")
    print("📊 Bulk CV upload with automatic ranking")
    print("✍️  Text input for job descriptions")
    print("📁 Excel database for data storage and reporting")
    print("\nEndpoints:")
    print("  GET  /                       - Web interface")
    print("  GET  /api/health             - Health check")
    print("  POST /api/upload-cvs-bulk    - Upload multiple CVs")
    print("  POST /api/upload-job-file    - Upload job document")
    print("  POST /api/add-job-text       - Add job as text")
    print("  POST /api/recommend          - Get ranked recommendations")
    print("  GET  /api/cvs                - List all CVs")
    print("  GET  /api/jobs               - List all jobs")
    print("  GET  /api/database-status    - Check Excel database status")
    print("  POST /api/clear              - Clear all data")
    print("\n📂 Excel Database Location: database/")
    print("   - cvs_database.xlsx")
    print("   - jobs_database.xlsx")
    print("   - recommendations_database.xlsx")
    print("\n" + "="*70)
    print(f"\n🌐 Server starting at: http://localhost:5000\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)